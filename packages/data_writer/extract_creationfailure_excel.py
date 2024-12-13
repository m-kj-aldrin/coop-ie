import locale
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
import os

locale.setlocale(locale.LC_TIME, "sv_SE.UTF-8")


def extract_key_values(text):
    """
    Extracts specific key-value pairs from the provided text based on predefined key sets
    and maps them to standardized headers.

    Parameters:
        text (str): The multiline text block containing key-value pairs.

    Returns:
        tuple: A dictionary containing the extracted key-value pairs mapped to standardized headers,
               and a boolean indicating whether the second key set was used.
    """
    # Define the two sets of keys
    first_type_keys = [
        "Error",
        "Pnr",
        "channel",
        "storeId",
        "Time",
        "receiptNumber",
        "id",
    ]
    second_type_keys = [
        "Till kundservice med beskrivning",
        "Personnummer",
        "Kanal",
        "Store id",
        "Ansökningsdatum",
        "Kvittonummer",
        "Id",
        "Epost",
    ]

    # Define the key indicator regex pattern
    key_indicator = "Till kundservice med beskrivning|Tekniskt fel eller fall som inte hanteras med beskrivning"
    indicator_pattern = f"^({key_indicator}): .*"
    indicator_regex = re.compile(indicator_pattern, re.MULTILINE | re.IGNORECASE)

    # Determine which set of keys to use
    use_second_type_keys = bool(indicator_regex.search(text))
    selected_keys = second_type_keys if use_second_type_keys else first_type_keys

    # Extract key-value pairs
    extracted = {}
    for key in selected_keys:
        escaped_key = re.escape(key)
        pattern = f"^{escaped_key}:(.*)$"
        regex = re.compile(pattern, re.MULTILINE | re.IGNORECASE)
        match = regex.search(text)
        value = match.group(1).strip() if match else ""
        extracted[key] = value

    # Handle special case for 'Personnummer' in second_type_keys
    if use_second_type_keys and "Personnummer" in selected_keys:
        personnummer = extracted.get("Personnummer", "")
        if personnummer:
            first_two_digits = personnummer[:2]
            if not personnummer.startswith(("19", "20")):
                if "20" <= first_two_digits <= "99":
                    extracted["Personnummer"] = "19" + personnummer
                elif "00" <= first_two_digits <= "20":
                    extracted["Personnummer"] = "20" + personnummer

    # Mapping to standardized headers
    headers = [
        "Orsak",
        "Personnummer",
        "Kanal",
        "Butiksnummer",
        "Ansökningsdatum",
        "Kvittonummer",
        "Ordernummer",
        "Epost",
    ]
    standardized_data = {header: "" for header in headers}

    if use_second_type_keys:
        key_mapping = {
            "Till kundservice med beskrivning": "Orsak",
            "Personnummer": "Personnummer",
            "Kanal": "Kanal",
            "Store id": "Butiksnummer",
            "Ansökningsdatum": "Ansökningsdatum",
            "Kvittonummer": "Kvittonummer",
            "Id": "Ordernummer",
            "Epost": "Epost",
        }
    else:
        key_mapping = {
            "Error": "Orsak",
            "Pnr": "Personnummer",
            "channel": "Kanal",
            "storeId": "Butiksnummer",
            "Time": "Ansökningsdatum",
            "receiptNumber": "Kvittonummer",
            "id": "Ordernummer",
            # 'Epost' remains blank
        }

    for original_key, value in extracted.items():
        standardized_key = key_mapping.get(original_key)
        if standardized_key:
            standardized_data[standardized_key] = value

    return standardized_data, use_second_type_keys


def determine_sheet_name(date_str):
    """
    Determines the sheet name based on the provided date string.

    Parameters:
        date_str (str): The date string extracted from the data.

    Returns:
        str: The sheet name in the format "YYYY MMM" (e.g., "2023 Nov").
    """
    if not date_str:
        # If no date is provided, default to current month
        return datetime.now().strftime("%Y %b")

    # Attempt to parse the date string with known formats
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%b %Y")  # e.g., "2023 Nov"
        except ValueError:
            continue
    # If parsing fails, default to current month
    return datetime.now().strftime("%b %Y")


def write_to_excel_append(output_file, data_dict):
    """
    Appends the extracted data along with CAS-nummer to the appropriate monthly sheet in the Excel file.

    Parameters:
        output_file (str): The path to the Excel file.
        data_dict (dict): Dictionary containing the extracted key-value pairs.
    """
    sheet_name = determine_sheet_name(data_dict.get("Ansökningsdatum", ""))

    # Define the headers in the desired order
    headers = [
        "Orsak",
        "Personnummer",
        "Kanal",
        "Butiksnummer",
        "Ansökningsdatum",
        "Kvittonummer",
        "Ordernummer",
        "Epost",
        "CAS-nummer",
    ]

    # Add 'CAS-nummer' to the data dictionary
    data_dict["CAS-nummer"] = data_dict.get("CAS-nummer", "")

    # Reorder the data_dict according to headers
    row = [data_dict.get(header, "") for header in headers]

    # Check if the Excel file exists
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
    else:
        wb = Workbook()
        # Remove the default sheet created by openpyxl
        default_sheet = wb.active
        wb.remove(default_sheet)

    # print(f"Sheet names in '{output_file}': {wb.sheetnames}")

    for name in wb.sheetnames:
        _name = name.lower()
        if _name == sheet_name.lower():
            ws = wb[name]
            break

    # if sheet_name in wb.sheetnames:
    #     ws = wb[sheet_name]
    #     print(f"Sheet '{sheet_name}' already exists in '{output_file}'. {ws}")
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)  # Write headers to the new sheet

    ws.append(row)  # Append the data row to the sheet
    wb.save(output_file)
    # print(f"Data appended to sheet '{sheet_name}' in '{output_file}'.")


def write_to_excel_multiple_append(data_list, cas_ids, output_file="output.xlsx"):
    """
    Writes multiple extracted data dictionaries along with CAS-nummer to an Excel file with specified headers.
    Each entry is appended to the appropriate monthly sheet based on 'Ansökningsdatum'.

    Parameters:
        data_list (list): List of dictionaries containing the extracted key-value pairs.
        cas_ids (list): List of CAS-nummer corresponding to each data dictionary.
        output_file (str): The name of the output Excel file.
    """
    for data_dict, cas_id in zip(data_list, cas_ids):
        # Add 'CAS-nummer' to the data dictionary
        data_dict["CAS-nummer"] = cas_id

        # If using first_type_keys, ensure 'Epost' is blank
        if "Epost" not in data_dict or not data_dict.get("Epost"):
            data_dict["Epost"] = ""

        write_to_excel_append(output_file, data_dict)


def main():
    # List of entries with text and corresponding cas_id
    entries = [
        {
            "text": """Mach 1 stötte på ett fel av typen: MembershipCreation

Error: Error obtaining primary address for customer - no primary address found
Time: 2024-12-12T14:57:25.655Z
Environment: PRODUCTION
receiptNumber: 300
storeId: 015480
entryCode: BUTPG
Pnr: 198706011841
date: 2024-12-12T00:00:00.000Z
channel: MIK
id: 86d257ee-c4fe-4cef-97c6-ac988dc5d720

Ta med referensnummret vid supportfrågor om detta ärende: f36854d2-ad15-4472-b101-b01e6838be70

Med vänliga hälsningar

The Mach1 Team
""",
            "cas_id": "CAS-8148288-X7N7Q6",
        },
    ]

    # Extract data from all entries
    data_list = []
    cas_ids = []
    for entry in entries:
        text = entry["text"]
        cas_id = entry["cas_id"]
        extracted_data, use_second_type_keys = extract_key_values(text)

        # If using first_type_keys, ensure 'Epost' is blank
        if not use_second_type_keys:
            extracted_data["Epost"] = ""

        data_list.append(extracted_data)
        cas_ids.append(cas_id)

    # Define the output Excel file path
    output_file = "app/data/creation_failure__.xlsx"

    # Write all entries to Excel
    write_to_excel_multiple_append(data_list, cas_ids, output_file=output_file)

    # For verification, print all extracted data
    print("\nExtracted Data Dictionaries:")
    headers = [
        "Orsak",
        "Personnummer",
        "Kanal",
        "Butiksnummer",
        "Ansökningsdatum",
        "Kvittonummer",
        "Ordernummer",
        "Epost",
        "CAS-nummer",
    ]
    for idx, (data, cas_id) in enumerate(zip(data_list, cas_ids), start=1):
        print(f"\nEntry {idx}:")
        for header in headers:
            if header != "CAS-nummer":
                print(f"{header}: {data.get(header, '')}")
            else:
                print(f"{header}: {cas_id}")


if __name__ == "__main__":
    main()
