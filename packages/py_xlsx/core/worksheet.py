"""Type-safe Excel worksheet wrapper using dataclasses and Pydantic models."""

from dataclasses import fields as dataclass_fields
from typing import Any, TypeVar, Generic, Iterator
from pydantic import BaseModel

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.workbook import Workbook


T = TypeVar("T")


class TypedWorkSheet(Generic[T]):
    """A wrapper around openpyxl Worksheet that provides type-safe operations.
    
    This class supports both dataclasses and Pydantic models.
    """

    worksheet: Worksheet
    _table_style: str = "TableStyleMedium9"

    def __init__(
        self,
        workbook: Workbook,
        model_type: type[T],
        sheet_name: str | None = None,
        table_name: str | None = None,
    ):
        self.workbook: Workbook = workbook
        self.model_type: type[T] = model_type
        self.sheet_name: str = sheet_name or model_type.__name__
        self.table_name: str | None = table_name or model_type.__name__
        
        # Handle both Pydantic models and dataclasses
        if issubclass(model_type, BaseModel):
            self._expected_headers = list(model_type.model_fields.keys())
        else:
            self._expected_headers = [field.name for field in dataclass_fields(model_type)]

        # Rest of initialization remains the same
        self._last_column_letter = chr(64 + len(self._expected_headers))
        
        if self.sheet_name not in workbook.sheetnames:
            self.worksheet = workbook.create_sheet(self.sheet_name)
            self._init_headers()
        else:
            self.worksheet = workbook[self.sheet_name]
            if not self._validate_headers():
                self._init_headers()

        if self.table_name and self.table_name not in self.worksheet.tables:
            self._create_table()

    def _init_headers(self) -> None:
        """Initialize the worksheet with headers from the dataclass."""
        for col, header in enumerate(self._expected_headers, 1):
            _ = self.worksheet.cell(row=1, column=col, value=header)

    def _validate_headers(self) -> bool:
        """Check if the worksheet headers match the dataclass fields."""
        if self.worksheet.max_row == 0:
            return False

        for col, expected_header in enumerate(self._expected_headers, 1):
            cell_value = self.worksheet.cell(row=1, column=col).value
            if cell_value != expected_header:
                return False
        return True

    def _create_table(self) -> None:
        """Create a table in the worksheet if it doesn't exist."""
        if not self.table_name or self.worksheet.max_row < 1:
            return

        tab = Table(
            displayName=self.table_name,
            ref=f"A1:{self._last_column_letter}{max(2, self.worksheet.max_row)}",
        )

        style = TableStyleInfo(
            name=self._table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style

        self.worksheet.add_table(tab)

    def _update_table_range(self) -> None:
        """Update the table range to include all rows."""
        if not self.table_name:
            return

        table: Table | None = self.worksheet.tables.get(self.table_name)
        if table and self.worksheet.max_row > 0:
            table.ref = f"A1:{self._last_column_letter}{self.worksheet.max_row}"

    def _dataclass_to_row(self, item: T) -> list[Any]:
        """Convert a dataclass or Pydantic model instance to a row format."""
        if isinstance(item, BaseModel):
            return [getattr(item, field) for field in self._expected_headers]
        else:
            return [getattr(item, field.name) for field in dataclass_fields(self.model_type)]

    def append(self, item: T) -> None:
        """Append a dataclass or Pydantic model instance to the worksheet."""
        if not isinstance(item, self.model_type):
            raise TypeError(
                f"Expected {self.model_type.__name__}, got {type(item).__name__}"
            )

        _ = self.worksheet.append(self._dataclass_to_row(item))
        
        if self.table_name and self.table_name not in self.worksheet.tables:
            self._create_table()
        else:
            self._update_table_range()

    def insert_row(self, item: T, row_index: int) -> None:
        """Insert a dataclass instance at a specific row in the worksheet.

        Args:
            item: An instance of the dataclass type specified in __init__
            row_index: The row number to insert at (1-based indexing, header is row 1)
        
        Raises:
            TypeError: If item is not of the expected dataclass type
            ValueError: If row_index is less than 2 (can't insert in header row)
        """
        if not isinstance(item, self.model_type):
            raise TypeError(
                f"Expected {self.model_type.__name__}, got {type(item).__name__}"
            )
        
        if row_index < 2:
            raise ValueError("Cannot insert into header row (row 1)")

        # Insert a blank row
        self.worksheet.insert_rows(row_index)
        
        # Convert dataclass to row data
        row_data = self._dataclass_to_row(item)
        
        # Fill in the cells
        for col, value in enumerate(row_data, 1):
            _ = self.worksheet.cell(row=row_index, column=col, value=value)
        
        # Update table range if needed
        if self.table_name in self.worksheet.tables:
            self._update_table_range()

    def delete_row(self, row_index: int) -> None:
        """Delete a specific row from the worksheet.

        Args:
            row_index: The row number to delete (1-based indexing, header is row 1)
        
        Raises:
            ValueError: If trying to delete header row or invalid row index
        """
        if row_index < 2:
            raise ValueError("Cannot delete header row (row 1)")
        
        if row_index > self.worksheet.max_row:
            raise ValueError(f"Row index {row_index} is out of range")

        self.worksheet.delete_rows(row_index)
        
        # Update table range if needed
        if self.table_name in self.worksheet.tables:
            self._update_table_range()

    def prepend(self, item: T) -> None:
        """Insert a dataclass instance at the beginning of the worksheet (after headers).

        Args:
            item: An instance of the dataclass type specified in __init__

        Raises:
            TypeError: If item is not of the expected dataclass type
        """
        self.insert_row(item, 2)  # Row 2 is first row after headers

    def iter_rows(self, skip_header: bool = True) -> Iterator[T]:
        """Iterate over worksheet rows as dataclass instances.
        
        Args:
            skip_header: Whether to skip the header row. Defaults to True.
            
        Yields:
            Instances of the dataclass type specified in __init__
        """
        start_row = 2 if skip_header else 1
        for row in self.worksheet.iter_rows(min_row=start_row, values_only=True):
            yield self.model_type(**dict(zip(self._expected_headers, row)))

    @property
    def row_count(self) -> int:
        """Return the number of data rows (excluding header)."""
        return max(0, self.worksheet.max_row - 1) 